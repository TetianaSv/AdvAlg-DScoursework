
import random
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from chained_hashtable import ChainedHashTable
from hash_functions import hashpjw

def build_station_table(stations, m=10):
    """Create a chained hash table containing all given station names."""
    table = ChainedHashTable(m=m, hash_func=hashpjw)
    for station in stations:
        table.insert(station)  # insert key
    return table

def check_station_status(table, name):
    """Return whether a station exists in the hash table."""
    return "Operational" if table.search(name) else "Non-operational"

# Pt 1 Performance Measure
def measure_lookup_time(n: int, num_queries: int = 1000) -> float:
    """Measure average lookup time for n inserted items."""
    stations = [f"Station_{i}" for i in range(n)]
    table = build_station_table(stations, m=n * 2)

    queries = random.sample(stations, num_queries)
    start = time.time()
    for q in queries:
        _ = table.search(q)
    end = time.time()
    return (end - start)/num_queries


dataset_sizes = [1000, 5000, 10000, 25000, 50000]
avg_times = []

print("Empirical performance test (Task 1b)\n")
for n in dataset_sizes:
    t = measure_lookup_time(n)
    avg_times.append(t)
    print(f"n = {n:<6} | Avg lookup time = {t:.8f} s")
#pt 2 result
plt.figure(figsize=(8, 5))
plt.plot(dataset_sizes, avg_times, marker="o", color="royalblue")
plt.title("Average Lookup Time vs Dataset Size n")
plt.xlabel("Number of Stations (n)")
plt.ylabel("Average Lookup Time (seconds)")
plt.grid(True)
plt.tight_layout()
plt.show()

print("\nTheoretical complexity: O(1) average lookup time.")
print("As number grows, measured times stay roughly constant .\n")

excel_file = "London Underground data.xlsx"
wb = load_workbook(excel_file)
ws = wb.active
# Collect all unique stations from the first two columns
stations_real = set()
for row in ws.iter_rows(values_only=True):
    if row[0]:
        stations_real.add(row[0])
    if row[1]:
        stations_real.add(row[1])

table_real = build_station_table(stations_real, m=len(stations_real) * 2)

print(f"Total unique stations loaded: {len(stations_real)}\n")

def test_real_stations():
    tests = ["Greenfort", "Paddington"]
    for name in tests:
        print(f"{name}: {check_station_status(table_real, name)}")

print("London Underground Operational Status Check...")
test_real_stations()
