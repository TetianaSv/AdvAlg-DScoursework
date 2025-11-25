#task 2b testing generated stations by aidanas alyta

import sys
import os
import random
import time

# library import
from adjacency_list_graph import AdjacencyListGraph
from dijkstra import dijkstra

# build a random connected undirected weighted graph
def build_random_connected_graph(n, avg_degree=6, w_min=1, w_max=10, rng=None):
    if rng is None:
        rng = random.Random()

    G = AdjacencyListGraph(n, directed=False, weighted=True)

    # connect each new station to a random earlier one to keep the graph connected
    for i in range(1, n):
        j = rng.randrange(0, i)
        w = rng.randint(w_min, w_max)
        G.insert_edge(j, i, w)

    # add extra random edges until we reach the desired density
    target_edges = int(n * avg_degree / 2.0)
    current_edges = G.get_card_E()
    attempts = 0

    while current_edges < target_edges and attempts < n * n:
        u = rng.randrange(0, n)
        v = rng.randrange(0, n)
        if u != v and not G.has_edge(u, v):
            w = rng.randint(w_min, w_max)
            try:
                G.insert_edge(u, v, w)
                current_edges += 1
            except RuntimeError:
                pass
        attempts += 1

    return G

# pick two different random station numbers
def pick_distinct_pair(n, rng):
    s = rng.randrange(0, n)
    t = rng.randrange(0, n - 1)
    if t >= s:
        t += 1
    return s, t

# turn predecessor list into an ordered path
def reconstruct_path(pi, s, t):
    path = []
    cur = t
    while cur is not None:
        path.insert(0, cur)
        if cur == s:
            break
        cur = pi[cur]
    return path

# main benchmarking function
def benchmark_dijkstra(ns, trials_per_n=200, avg_degree=6, w_min=1, w_max=10, seed=42):
    rng = random.Random(seed)

    print("=======================================================================")
    print("TASK 2b: Measuring Dijkstra's Performance on Random Tube Networks")
    print("=======================================================================")
    print("Average degree:", avg_degree, "| weight range:", w_min, "-", w_max,
          "| trials per network:", trials_per_n, "| seed:", seed)
    print("-----------------------------------------------------------------------")
    print("Columns: Stations, Edges, Average Time (ms)")
    print("-----------------------------------------------------------------------")

    for n in ns:
        # build random connected graph
        G = build_random_connected_graph(n, avg_degree=avg_degree,
                                         w_min=w_min, w_max=w_max, rng=rng)
        m = G.get_card_E()

        # measure Dijkstra average time
        times = []
        for _ in range(trials_per_n):
            s, t = pick_distinct_pair(n, rng)
            t_start = time.perf_counter()
            d, pi = dijkstra(G, s)
            t_end = time.perf_counter()
            elapsed_ms = (t_end - t_start) * 1000.0
            times.append(elapsed_ms)

        avg_ms = sum(times) / float(len(times))

        # print average results
        print(str(n) + " , " + str(m) + " , " + str(round(avg_ms, 3)))


# test with different amount of stations
if __name__ == "__main__":
    sizes = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]

    benchmark_dijkstra(
        ns=sizes,
        trials_per_n=200,
        avg_degree=6,
        w_min=1,
        w_max=10,
        seed=42
    )

#part 2b:London underground data import
from adjacency_list_graph import AdjacencyListGraph
from dijkstra import dijkstra

#required import library to open excel file
from openpyxl import load_workbook

def load_edges_from_excel(file_path):

    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active  # first sheet
    edges = []
    skipped = 0
    dupes = 0

    first_row = True
    for row in ws.iter_rows(values_only=True):
        # Skip header row
        if first_row:
            first_row = False
            continue

        if row is None or len(row) < 4:
            skipped += 1
            continue

        src = str(row[1]).strip() if row[1] else ""
        dst = str(row[2]).strip() if row[2] else ""
        mins = row[3]

        if not src or not dst or mins is None:
            skipped += 1
            continue
        try:
            w = float(mins)
        except ValueError:
            skipped += 1
            continue
        edges.append((src, dst, w)) #store the edges

    # build station dictionary and graph 
    stations = []
    seen = set()
    for u, v, _ in edges:
        if u not in seen:
            seen.add(u)
            stations.append(u)
        if v not in seen:
            seen.add(v)
            stations.append(v)

    id_by_name = {name.upper(): i for i, name in enumerate(stations)}
    name_by_id = {i: name for i, name in enumerate(stations)}

    G = AdjacencyListGraph(len(stations), directed=False, weighted=True)

    for u_name, v_name, w in edges:
        u = id_by_name[u_name.upper()]
        v = id_by_name[v_name.upper()]
        if not G.has_edge(u, v):
            G.insert_edge(u, v, w)
        else:
            dupes += 1

    return G, id_by_name, name_by_id, skipped, dupes

#rebuild path from Dijkstra’s output
def reconstruct_path(pi, s, t):
    path = []
    cur = t
    while cur is not None:
        path.insert(0, cur)
        if cur == s:
            break
        cur = pi[cur]
    return path

# locate the excel file:
file_path = 'London Underground data.xlsx'

#main function
def main():

    print("Loading:", file_path)
    G, id_by_name, name_by_id, skipped, dupes = load_edges_from_excel(file_path)

    print("Stations:", G.get_card_V(), " | Connections:", G.get_card_E())
    print("Skipped rows:", skipped, " | Duplicates ignored:", dupes)
    print("")
    print("Type two station names to find the fastest route (blank to quit).")

    while True:
        src = input("Start station: ").strip()
        if src == "":
            break
        dst = input("Destination station: ").strip()
        if dst == "":
            break

        s_key = src.upper()
        t_key = dst.upper()

        # Make sure both stations exist
        if s_key not in id_by_name or t_key not in id_by_name:
            print("Unknown station name.")
            continue
        if s_key == t_key:
            print("Start and destination are the same.")
            continue

        s = id_by_name[s_key]
        t = id_by_name[t_key]

        # Run Dijkstra’s algorithm to find shortest travel times
        d, pi = dijkstra(G, s)

        # Reconstruct route from station IDs
        path_ids = reconstruct_path(pi, s, t)

        if len(path_ids) == 0 or d[t] == float('inf'):
            print("No path found.")
        else:
            path_names = [name_by_id[i] for i in path_ids]
            print("Route:", " -> ".join(path_names))
            print("Total time:", str(d[t]), "minutes")
        print("")

if __name__ == "__main__":
    main()
